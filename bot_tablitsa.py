import re
import logging
from collections import OrderedDict
from pathlib import Path

import telebot
from telebot import types
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.formatting.rule import CellIsRule

# ====== Настройки ======
from secret import TOKEN  # ваш токен
TELEGRAM_TOKEN = TOKEN
XLSX_PATH = Path(__file__).parent / "Асортимент.xlsx"
SHEET_NAME = "Лист1"

bot = telebot.TeleBot(TELEGRAM_TOKEN)
logging.basicConfig(level=logging.INFO)

# ====== Хранилище состояний пользователей ======
user_data = {}  # chat_id → dict: action, type, category, row, name, stock, awaiting_qty

# ====== Регулярка для "X шт" ======
qty_re = re.compile(r"(\d+)")

# ====== Заливки ======
RED_FILL = PatternFill(fill_type="solid", fgColor=Color(rgb="FFFF0000"))
GREEN_FILL = PatternFill(fill_type="solid", fgColor=Color(theme=9, tint=0.4))

def apply_conditional_formatting(sheet):
    """
    Добавляет два правила условного форматирования для столбца C:
      - равно "0 шт" → красная заливка
      - не равно "0 шт" → зелёная заливка (Accent6, +40%)
    Очищает старые правила, чтобы не дублировать.
    """
    rng = f"C4:C{sheet.max_row}"
    red_rule = CellIsRule(operator='equal',    formula=['"0 шт"'],   fill=RED_FILL)
    green_rule = CellIsRule(operator='notEqual', formula=['"0 шт"'], fill=GREEN_FILL)

    # Сбрасываем все старые правила
    sheet.conditional_formatting._cf_rules.clear()
    sheet.conditional_formatting.add(rng, red_rule)
    sheet.conditional_formatting.add(rng, green_rule)

def load_sheet():
    """Загружает Excel, применяет CF и возвращает (wb, sheet)."""
    wb = load_workbook(XLSX_PATH)
    sheet = wb[SHEET_NAME]
    apply_conditional_formatting(sheet)
    return wb, sheet

def build_catalog():
    """
    Сканирует лист и возвращает структуру:
      {
        "Жидкости": OrderedDict({
          "Рик и Морти на замерзоне": [(name, row), …],
          "Catswill": [(name, row), …]
        }),
        "Картриджи": OrderedDict({
          "Расходники": [(name, row), …]
        })
      }
    """
    wb, sheet = load_sheet()
    catalog = {"Жидкости": OrderedDict(), "Картриджи": OrderedDict()}
    current_type = current_cat = None

    for r in range(1, sheet.max_row + 1):
        name  = sheet.cell(r, 1).value
        vol   = sheet.cell(r, 2).value
        stock = sheet.cell(r, 3).value

        if name in ("CLOUD HAVEN", "Catswill"):
            cat = "Рик и Морти на замерзоне" if name == "CLOUD HAVEN" else name
            catalog["Жидкости"][cat] = []
            current_type, current_cat = "Жидкости", cat

        elif name == "Расходники":
            catalog["Картриджи"][name] = []
            current_type, current_cat = "Картриджи", name

        elif name and vol and stock and current_cat:
            catalog[current_type][current_cat].append((name, r))

    return catalog

# ====== Шаг 0: /start и выбор действия ======
@bot.message_handler(commands=['start'])
def cmd_start(message):
    chat_id = message.chat.id
    user_data.pop(chat_id, None)
    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton("Продать",   callback_data="action:sell"),
        types.InlineKeyboardButton("Пополнить", callback_data="action:add")
    )
    bot.send_message(chat_id, "Выберите действие:", reply_markup=kb)

# ====== Шаг 1: выбор действия ======
@bot.callback_query_handler(func=lambda c: c.data.startswith("action:"))
def handle_action(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    action  = call.data.split("action:")[1]
    user_data[chat_id] = {"action": action}
    bot.answer_callback_query(call.id)

    prompt = "Выберите тип для продажи:" if action == "sell" else "Выберите тип для пополнения:"
    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton("Жидкости",   callback_data="type:Жидкости"),
        types.InlineKeyboardButton("Картриджи", callback_data="type:Картриджи")
    )
    bot.edit_message_text(prompt, chat_id, call.message.message_id, reply_markup=kb)

# ====== Шаг 2: выбор типа товара ======
@bot.callback_query_handler(func=lambda c: c.data.startswith("type:"))
def handle_type(call: types.CallbackQuery):
    chat_id  = call.message.chat.id
    sel_type = call.data.split("type:")[1]
    user_data[chat_id]["type"] = sel_type
    bot.answer_callback_query(call.id)

    cats = build_catalog().get(sel_type, {})
    action = user_data[chat_id]["action"]
    prompt = f"{'Продажа' if action=='sell' else 'Пополнение'} → {sel_type}.\nВыберите подкатегорию:"
    kb = types.InlineKeyboardMarkup()
    for cat in cats.keys():
        kb.add(types.InlineKeyboardButton(cat, callback_data=f"cat:{cat}"))
    bot.edit_message_text(prompt, chat_id, call.message.message_id, reply_markup=kb)

# ====== Шаг 3: выбор подкатегории ======
@bot.callback_query_handler(func=lambda c: c.data.startswith("cat:"))
def handle_category(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    sel_cat = call.data.split("cat:")[1]
    user_data[chat_id]["category"] = sel_cat
    bot.answer_callback_query(call.id)

    sel_type = user_data[chat_id]["type"]
    products = build_catalog()[sel_type][sel_cat]
    action = user_data[chat_id]["action"]
    prompt = f"{'Продажа' if action=='sell' else 'Пополнение'} → {sel_type} → {sel_cat}.\nВыберите товар:"
    kb = types.InlineKeyboardMarkup()
    for name, row in products:
        kb.add(types.InlineKeyboardButton(name, callback_data=f"prod:{row}"))
    bot.edit_message_text(prompt, chat_id, call.message.message_id, reply_markup=kb)

# ====== Шаг 4: выбор товара ======
@bot.callback_query_handler(func=lambda c: c.data.startswith("prod:"))
def handle_product(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    row     = int(call.data.split("prod:")[1])
    wb, sheet = load_sheet()
    name      = sheet.cell(row, 1).value
    stock_txt = sheet.cell(row, 3).value or "0"
    current   = int(qty_re.search(stock_txt).group(1))

    user_data[chat_id].update({"row": row, "name": name, "stock": current})
    bot.answer_callback_query(call.id)

    action = user_data[chat_id]["action"]
    prompt = (
        f'Товар: "{name}". Остаток: {current} шт.\n'
        + ("Сколько продать?" if action == "sell" else "Сколько добавить?")
    )
    kb = types.InlineKeyboardMarkup()
    for n in (1, 2, 5):
        kb.add(types.InlineKeyboardButton(str(n), callback_data=f"qty:{n}"))
    kb.add(types.InlineKeyboardButton("Другое", callback_data="qty:0"))

    bot.edit_message_text(prompt, chat_id, call.message.message_id, reply_markup=kb)

# ====== Шаг 5: выбор/ввод количества ======
@bot.callback_query_handler(func=lambda c: c.data.startswith("qty:"))
def handle_qty(call: types.CallbackQuery):
    chat_id = call.message.chat.id
    qty     = int(call.data.split("qty:")[1])
    data    = user_data.get(chat_id, {})
    bot.answer_callback_query(call.id)

    if qty == 0:
        data["awaiting_qty"] = True
        bot.edit_message_text("Пожалуйста, введите число:", chat_id, call.message.message_id)
        return

    _finalize(chat_id, qty)

@bot.message_handler(func=lambda m: user_data.get(m.chat.id, {}).get("awaiting_qty"))
def handle_manual_qty(message: types.Message):
    chat_id = message.chat.id
    text    = message.text.strip()
    if not text.isdigit():
        return bot.reply_to(message, "Введите число.")
    qty = int(text)
    user_data[chat_id].pop("awaiting_qty", None)
    _finalize(chat_id, qty)

def _finalize(chat_id: int, qty: int):
    """
    Общая финализация для продажи или пополнения:
      - корректировка остатка
      - запись в Excel + мануальная заливка
      - отправка итогового сообщения
    """
    data    = user_data[chat_id]
    action  = data["action"]
    row     = data["row"]
    name    = data["name"]
    current = data["stock"]
    new     = (current - qty) if action == "sell" else (current + qty)

    if action == "sell" and new < 0:
        bot.send_message(chat_id, f"Недостаточно на складе ({current} шт).")
        return

    wb, sheet = load_sheet()
    cell = sheet.cell(row, 3)
    cell.value = f"{new} шт"
    cell.fill  = RED_FILL if new == 0 else GREEN_FILL
    wb.save(XLSX_PATH)

    verb = "Продано" if action == "sell" else "Добавлено"
    bot.send_message(chat_id, f'{verb} {qty} шт "{name}". Новый остаток: {new} шт.')
    user_data.pop(chat_id, None)

# ====== Запуск бота ======
if __name__ == "__main__":
    print("Бот запущен...")
    bot.infinity_polling()
